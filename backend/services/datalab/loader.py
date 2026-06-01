"""
DataLab loader — turn raw spreadsheet/CSV bytes into an in-memory table set.

Fidelity-first: every cell is read as a STRING. This guarantees zero silent
coercion — leading zeros (ZIP codes), long numeric IDs, and dates-as-written
survive exactly. Numeric reasoning happens later via explicit DuckDB TRY_CAST,
never by mutating the source on ingest.

Returns a `LoadedWorkbook`:
  .sheets : list[Sheet]   (CSV/TSV => single sheet named "Sheet1")
  Sheet:
    .name      str
    .columns   list[str]
    .rows      list[list[str]]   (ALL rows; empty cell => "")
    .row_count int
    .encoding  str | None        (CSV only)
    .delimiter str | None        (CSV only)
"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from typing import List, Optional

from .config import MAX_SHEETS


@dataclass
class Sheet:
    name: str
    columns: List[str]
    rows: List[List[str]]
    encoding: Optional[str] = None
    delimiter: Optional[str] = None

    @property
    def row_count(self) -> int:
        return len(self.rows)


@dataclass
class LoadedWorkbook:
    kind: str                       # 'xlsx' | 'xls' | 'csv' | 'tsv'
    sheets: List[Sheet] = field(default_factory=list)

    @property
    def total_rows(self) -> int:
        return sum(s.row_count for s in self.sheets)


class LoadError(ValueError):
    """Raised when bytes cannot be parsed as a spreadsheet/CSV."""


def _detect_encoding(raw: bytes) -> str:
    """Best-effort charset detection with a safe utf-8 fallback."""
    # BOM checks first (cheap + authoritative)
    if raw.startswith(b"\xff\xfe") or raw.startswith(b"\xfe\xff"):
        return "utf-16"
    if raw.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"
    try:
        from charset_normalizer import from_bytes
        match = from_bytes(raw).best()
        if match and match.encoding:
            return match.encoding
    except Exception:
        pass
    return "utf-8"


def _detect_delimiter(sample_text: str, filename: str) -> str:
    """Sniff CSV delimiter; honor .tsv; fall back to comma."""
    if filename.lower().endswith(".tsv"):
        return "\t"
    try:
        dialect = csv.Sniffer().sniff(sample_text, delimiters=",;\t|")
        return dialect.delimiter
    except Exception:
        # Heuristic fallback: pick the most common candidate in the header line
        head = sample_text.splitlines()[0] if sample_text.splitlines() else ""
        counts = {d: head.count(d) for d in [",", ";", "\t", "|"]}
        best = max(counts, key=counts.get)
        return best if counts[best] > 0 else ","


def _load_csv(raw: bytes, filename: str) -> LoadedWorkbook:
    enc = _detect_encoding(raw)
    try:
        text = raw.decode(enc, errors="replace")
    except LookupError:
        enc = "utf-8"
        text = raw.decode(enc, errors="replace")
    # Strip a leading BOM char if the codec left one
    if text and text[0] == "\ufeff":
        text = text[1:]
    sample = text[:8192]
    delim = _detect_delimiter(sample, filename)

    reader = csv.reader(io.StringIO(text), delimiter=delim)
    all_rows = [list(r) for r in reader]
    if not all_rows:
        raise LoadError("empty CSV")
    columns = [c.strip() for c in all_rows[0]]
    width = len(columns)
    body: List[List[str]] = []
    for r in all_rows[1:]:
        # normalize ragged rows to header width (pad short, keep extra)
        if len(r) < width:
            r = r + [""] * (width - len(r))
        body.append([("" if c is None else str(c)) for c in r])
    kind = "tsv" if delim == "\t" else "csv"
    sheet = Sheet(name="Sheet1", columns=columns, rows=body,
                  encoding=enc, delimiter=delim)
    return LoadedWorkbook(kind=kind, sheets=[sheet])


def _load_excel(raw: bytes, kind: str) -> LoadedWorkbook:
    import openpyxl  # only needed for xlsx
    if kind == "xls":
        # Legacy .xls: fall back to pandas+xlrd if available
        try:
            import pandas as pd
            dfs = pd.read_excel(io.BytesIO(raw), sheet_name=None, dtype=str,
                                keep_default_na=False)
        except Exception as e:
            raise LoadError(f"cannot read .xls: {e}")
        sheets = []
        for name, df in list(dfs.items())[:MAX_SHEETS]:
            cols = [str(c) for c in df.columns]
            rows = df.astype(str).values.tolist()
            sheets.append(Sheet(name=str(name), columns=cols, rows=rows))
        return LoadedWorkbook(kind=kind, sheets=sheets)

    wb = openpyxl.load_workbook(
        io.BytesIO(raw), read_only=True, data_only=True, keep_links=False
    )
    sheets: List[Sheet] = []
    for ws in wb.worksheets[:MAX_SHEETS]:
        it = ws.iter_rows(values_only=True)
        try:
            header = next(it)
        except StopIteration:
            sheets.append(Sheet(name=ws.title, columns=[], rows=[]))
            continue
        columns = [("" if c is None else str(c)) for c in header]
        width = len(columns)
        rows: List[List[str]] = []
        for r in it:
            vals = [("" if c is None else str(c)) for c in r]
            if len(vals) < width:
                vals += [""] * (width - len(vals))
            rows.append(vals)
        sheets.append(Sheet(name=ws.title, columns=columns, rows=rows))
    wb.close()
    if not sheets:
        raise LoadError("workbook has no sheets")
    return LoadedWorkbook(kind=kind, sheets=sheets)


def load_workbook(raw: bytes, filename: str, mime: str = "") -> LoadedWorkbook:
    """Entry point: dispatch on extension/mime and return a LoadedWorkbook."""
    name = (filename or "").lower()
    if name.endswith(".csv") or name.endswith(".tsv") or \
            (mime or "").startswith("text/csv") or \
            (mime or "").startswith("text/tab"):
        return _load_csv(raw, filename)
    if name.endswith(".xls"):
        return _load_excel(raw, "xls")
    # default to xlsx/xlsm
    return _load_excel(raw, "xlsx")
